from datetime import date
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from scd_report.aggregator import build_report, read_personal_workbook


def make_fixture(path: Path) -> None:
    wb = Workbook()
    personal = wb.active
    personal.title = "개인일일업무"
    personal["B4"], personal["C4"] = "일자", "이름"
    personal["B6"], personal["C6"] = date(2026, 7, 15), "홍길동"
    personal["C7"], personal["D7"], personal["E7"], personal["F7"] = "구분", "업무 명", "시간", "업무"
    personal["C8"], personal["D8"], personal["E8"], personal["F8"] = "문서 업무", "SOP, 규정", 2, "SOP 개정"
    personal["Q7"], personal["R7"], personal["S7"], personal["T7"] = "구분", "업무 명", "시간", "업무"
    personal["Q8"], personal["R8"], personal["S8"], personal["T8"] = "US Project", "US PO", 3, "PO 수행"
    team = wb.create_sheet("팀 일일업무")
    team["B1"] = date(2026, 1, 1)
    daily = wb.create_sheet("일일공수")
    daily["F4"], daily["AA4"] = "SOP, 규정", "US PO"
    daily["A5"], daily["B5"] = date(2026, 7, 15), 0
    daily["C5"], daily["D5"] = "=B5*8", "=SUM(E5:AP5)"
    weekly = wb.create_sheet("주간공수")
    weekly["A1"] = "기존 주간 서식"
    report = wb.create_sheet("주간보고")
    report["A1"] = "기존 보고 서식"
    wb.save(path)


class AggregatorTests(unittest.TestCase):
    def test_reads_personal_block(self):
        with tempfile.TemporaryDirectory() as directory:
            sample = Path(directory) / "sample.xlsx"
            make_fixture(sample)
            records = read_personal_workbook(sample)
            self.assertEqual({r.name for r in records}, {"홍길동"})
            self.assertEqual({r.work_date for r in records}, {date(2026, 7, 15)})
            self.assertTrue(any(r.task == "SOP, 규정" and r.hours == 2 for r in records))

    def test_build_preserves_sheets_and_adds_outputs(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            make_fixture(template)
            team = root / "team"
            team.mkdir()
            make_fixture(team / "member.xlsx")
            output = root / "result.xlsx"
            count, warnings = build_report(team, template, output)
            self.assertEqual(count, 2)
            self.assertEqual(warnings, [])
            wb = load_workbook(output, data_only=False)
            self.assertIn("자동취합_원본", wb.sheetnames)
            self.assertIn("주간보고 리스트", wb.sheetnames)
            self.assertEqual(wb["팀 일일업무"]["B1"].value.date(), date(2026, 7, 15))
            self.assertEqual(wb["일일공수"]["B5"].value, 1)
            self.assertEqual(wb["일일공수"]["F5"].value, 2)
            self.assertEqual(wb["일일공수"]["AA5"].value, 3)
            self.assertEqual(wb["주간공수"]["A1"].value, "기존 주간 서식")


if __name__ == "__main__":
    unittest.main()
