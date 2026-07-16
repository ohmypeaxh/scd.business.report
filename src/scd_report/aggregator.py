from __future__ import annotations

from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill


PERSONAL_SHEET = "개인일일업무"
RAW_SHEET = "자동취합_원본"
WEEKLY_LIST_SHEET = "주간보고 리스트"
DAILY_TEAM_SHEET = "팀 일일업무"
DAILY_HOURS_SHEET = "일일공수"


@dataclass(frozen=True)
class WorkRecord:
    work_date: date
    name: str
    category: str
    task: str
    hours: float
    details: str
    source_file: str


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _clean_task(value: object) -> str:
    return re.sub(r"\s+", " ", _text(value)).strip()


def _category_rows(ws, start: int, end: int, category_col: int, task_col: int,
                   hours_col: int, detail_col: int, work_date: date, name: str,
                   source: str) -> list[WorkRecord]:
    records: list[WorkRecord] = []
    current_category = ""
    for row in range(start, end + 1):
        if _text(ws.cell(row, category_col).value):
            current_category = _clean_task(ws.cell(row, category_col).value)
        task = _clean_task(ws.cell(row, task_col).value)
        hours = _number(ws.cell(row, hours_col).value)
        details = _text(ws.cell(row, detail_col).value)
        if task and (hours or details):
            records.append(WorkRecord(work_date, name, current_category, task,
                                      hours, details, source))
    return records


def read_personal_workbook(path: Path) -> list[WorkRecord]:
    """Read every repeated personal-work block in an input workbook."""
    wb = load_workbook(path, data_only=True, read_only=False)
    if PERSONAL_SHEET not in wb.sheetnames:
        return []
    ws = wb[PERSONAL_SHEET]
    records: list[WorkRecord] = []
    header_rows = [r for r in range(1, ws.max_row + 1)
                   if _text(ws.cell(r, 2).value) == "일자"
                   and _text(ws.cell(r, 3).value) == "이름"]
    for header in header_rows:
        value_row = header + 2
        work_date = _date(ws.cell(value_row, 2).value)
        name = _text(ws.cell(value_row, 3).value)
        if not work_date or not name:
            continue
        table_header = header + 3
        block_end = min(header + 27, ws.max_row)
        records.extend(_category_rows(ws, table_header + 1, block_end, 3, 4, 5, 6,
                                      work_date, name, path.name))
        records.extend(_category_rows(ws, table_header + 1, block_end, 17, 18, 19, 20,
                                      work_date, name, path.name))
    return records


def collect_records(folder: Path, template_path: Path) -> tuple[list[WorkRecord], list[str]]:
    records: list[WorkRecord] = []
    warnings: list[str] = []
    template_resolved = template_path.resolve()
    for path in sorted(folder.glob("*.xlsx")):
        if path.name.startswith("~$") or path.resolve() == template_resolved:
            continue
        try:
            found = read_personal_workbook(path)
            if found:
                records.extend(found)
            else:
                warnings.append(f"{path.name}: '{PERSONAL_SHEET}' 데이터 없음")
        except Exception as exc:  # one bad team file must not abort the whole run
            warnings.append(f"{path.name}: 읽기 실패 ({exc})")
    return records, warnings


def _sheet_headers(ws, row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _clean_task(ws.cell(row, col).value)
        if key:
            result[key] = col
    return result


DAILY_COLUMNS = {
    "SOP, 규정": 6, "BMR": 7, "교육": 8, "URS": 9, "기안": 10,
    "QRM": 12, "기타문서": 13, "결재업무": 14, "US회의": 24,
    "내부회의": 15, "생산 공정": 16, "생산부 전산 업무": 17,
    "쉬핑라벨 업무": 18, "BMR 발행 및 관리": 19, "소모품관리": 20,
    "로그북 관련": 21, "입고전표 업무": 22, "기타": 23,
    "US 행정 업무": 25, "US 문서 업무": 26, "US PO": 27, "US PV": 28,
    "US 장비 관련": 29, "내수 PO": 30, "내수 PV": 31, "문서 업무": 32,
    "FAT 외근": 33, "설치": 34, "작동관련": 35, "지원감(-)": 36,
    "지원옴(+)": 37, "시간내 Loss": 38, "US관련 잔업": 39,
    "US관련 특근": 40, "잔업": 41, "특근": 42, "시간외 기타": 43,
    "휴가": 44, "외근/출장": 45,
}


def _write_daily_hours(wb, records: Iterable[WorkRecord]) -> None:
    if DAILY_HOURS_SHEET not in wb.sheetnames:
        return
    ws = wb[DAILY_HOURS_SHEET]
    dates = {d: row for row in range(5, ws.max_row + 1)
             if (d := _date(ws.cell(row, 1).value))}
    grouped: dict[tuple[date, int], float] = defaultdict(float)
    people: dict[date, set[str]] = defaultdict(set)
    for record in records:
        target_col = DAILY_COLUMNS.get(_clean_task(record.task))
        if target_col:
            grouped[(record.work_date, target_col)] += record.hours
        people[record.work_date].add(record.name)
    for work_date, names in people.items():
        row = dates.get(work_date)
        if not row:
            continue
        ws.cell(row, 2).value = len(names)
        for col in sorted(set(DAILY_COLUMNS.values())):
            if ws.cell(row, col).data_type != "f":
                ws.cell(row, col).value = grouped.get((work_date, col)) or None


def _joined_details(records: Iterable[WorkRecord], categories: set[str] | None = None) -> str:
    lines: list[str] = []
    for record in records:
        if categories and record.category not in categories:
            continue
        detail = record.details.strip()
        if detail:
            for line in detail.splitlines():
                line = line.strip()
                if line and line not in lines:
                    lines.append(line)
    return "\n".join(lines)


def _write_team_daily(wb, records: list[WorkRecord]) -> None:
    if DAILY_TEAM_SHEET not in wb.sheetnames or not records:
        return
    ws = wb[DAILY_TEAM_SHEET]
    latest = max(r.work_date for r in records)
    todays = [r for r in records if r.work_date == latest]
    ws["B1"] = latest
    ws["B3"] = _joined_details(todays, {"문서 업무", "회의", "공정개발 업무"})
    overtime = [r for r in todays if r.task in {"US관련 잔업", "US관련 특근", "잔업", "특근"}]
    ws["B15"] = "\n".join(sorted({r.name for r in overtime}))
    ws["C15"] = sum(r.hours for r in overtime) or None
    ws["E14"] = _joined_details(overtime)
    ws["B18"] = _joined_details(todays, {"장비"})
    ws["B30"] = _joined_details(todays, {"US Project", "내수 신제품"})


def _style_new_sheet(ws, widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width


def _write_record_sheet(wb, title: str, records: list[WorkRecord]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(["일자", "이름", "구분", "업무 명", "시간", "업무 내용", "원본 파일"])
    for record in sorted(records, key=lambda r: (r.work_date, r.name, r.category, r.task)):
        ws.append([record.work_date, record.name, record.category, record.task,
                   record.hours or None, record.details, record.source_file])
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _style_new_sheet(ws, [12, 12, 18, 24, 10, 80, 28])


def build_report(team_folder: Path, template_path: Path, output_path: Path) -> tuple[int, list[str]]:
    records, warnings = collect_records(team_folder, template_path)
    if not records:
        raise ValueError("취합할 개인 일일업무 데이터가 없습니다.")
    wb = load_workbook(template_path, data_only=False, keep_links=True)
    _write_daily_hours(wb, records)
    _write_team_daily(wb, records)
    _write_record_sheet(wb, RAW_SHEET, records)
    _write_record_sheet(wb, WEEKLY_LIST_SHEET, records)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(records), warnings
